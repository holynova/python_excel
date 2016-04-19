# -*- coding: utf-8 -*-
from openpyxl import Workbook,load_workbook
import datetime,os
wb = Workbook()
max_pack = 10
print u"input max package number:"
max_pack = int(raw_input())
##################################################
for i in range(max_pack+2):
	wb.create_sheet()
cnt = -2
for sheet in wb.worksheets:
	sheet.title = u"包" + str(cnt)
	cnt += 1
wb.worksheets[0].title = u'汇总1'
wb.worksheets[1].title = u'汇总2'
wb.worksheets[2].title = u'中标结果'

##################################################
for sheet in wb.worksheets:
	print sheet.title +" ",


now = datetime.datetime.now().strftime('%y%m%d %H-%M-%S')
file_name = os.path.dirname(os.path.abspath(__file__))+'\\python_excel' + now + ".xlsx"
# wb.save(file_name)
print file_name +" is ready."



# wb = load_workbook('E:\\110HV.xlsx')


# ws_test = wb.create_sheet()
# ws_test.title = 'test'

# for sheet in wb:
# 	print sheet.title
# print '------------'

# ws_data = wb['data']


# print ws_data['a1']
# for i in range(10):
# 	print ws_data.cell(row = 1 ,column = i+1).value
# 	ws_test.cell(row = i+1,column = i+1).value = i
# wb.save('E:\\110HV.xlsx')
