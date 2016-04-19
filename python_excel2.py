# -*- coding: utf-8 -*-
from openpyxl import Workbook,load_workbook
# import datetime,os
# wb = Workbook()
# max_pack = 10
# print u"input max package number:"
# max_pack = int(raw_input())
# ##################################################
# for i in range(max_pack+2):
# 	wb.create_sheet()
# cnt = -2
# for sheet in wb.worksheets:
# 	sheet.title = u"包" + str(cnt)
# 	cnt += 1
# wb.worksheets[0].title = u'汇总1'
# wb.worksheets[1].title = u'汇总2'
# wb.worksheets[2].title = u'中标结果'

# ##################################################
# for sheet in wb.worksheets:
# 	print sheet.title +" ",


# now = datetime.datetime.now().strftime('%y%m%d %H-%M-%S')
# file_name = os.path.dirname(os.path.abspath(__file__))+'\\python_excel' + now + ".xlsx"
# # wb.save(file_name)
# print file_name +" is ready."



wb = load_workbook('E:\\test\HV.xlsx')
sht_data = wb['data']
print 'length = ',len(wb.worksheets)
max_row = 1
number = sht_data.cell(row = max_row,column =1).value
while number:
	# print number,
	max_row += 1
	number = sht_data.cell(row = max_row,column =1).value
print 'max_row = ' +str(max_row-1)


arr_province = [u"安徽",u"北京",u"福建",u"甘肃",u"国网节能服务",u"河北",u"河南",u"黑龙江",u"湖北",u"湖南",u"华北",u"吉林",u"冀北",u"江苏",u"江西",u"辽宁",u"鲁能",u"内蒙",u"内蒙古东部",u"宁夏",u"青海",u"山东",u"山西",u"陕西",u"上海",u"四川",u"天津",u"西藏",u"新疆",u"浙江",u"中国技术装备",u"重庆",]

# col_province = sht_data['E:E']
for row in sht_data.iter_rows("e2:e"+str(max_row-1)):
	for cell in row:
		if cell.value.find(u"山东") != -1:
			print cell.value," ", 
